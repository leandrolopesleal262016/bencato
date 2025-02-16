from flask import Flask, render_template, jsonify, request, redirect, url_for, send_file
import pandas as pd
import numpy as np
from datetime import datetime
import os
from werkzeug.utils import secure_filename
from io import BytesIO
import chardet
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

UPLOAD_FOLDER = 'uploads'
ALLOWED_EXTENSIONS = {'csv'}

app = Flask(__name__)
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER

# Cria a pasta de uploads se não existir
if not os.path.exists(UPLOAD_FOLDER):
    os.makedirs(UPLOAD_FOLDER)

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def clean_numeric(value):
    if pd.isna(value) or value == '-' or value == '':
        return 0
    if isinstance(value, str):
        return float(value.replace('.', '').replace(',', '.').strip())
    return float(value)

def convert_series_to_dict(series):
    if isinstance(series, pd.Series):
        return series.to_dict()
    return series

def load_data(filename):
    filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
    df = pd.read_csv(filepath, 
                     encoding='cp1252',
                     sep=';',
                     skiprows=6)
    
    df = df.loc[:, df.columns.notna()]
    
    colunas = list(df.columns)
    coluna_total = colunas[21] if len(colunas) > 21 else 'Total'
    
    column_mapping = {
        'Dt. entrada': 'data_entrada',
        'Cliente': 'cliente',
        'Obra/Centro de custo': 'obra',
        'Doc. financ.': 'documento',
        'Fornecedor': 'fornecedor',
        'Tipo Custo': 'tipo_custo',
        'Qtd.': 'quantidade',
        'Und.': 'unidade',
        'Insumo': 'insumo',
        'Fase': 'fase'
    }
    column_mapping[coluna_total] = 'total'
    
    colunas_existentes = [col for col in column_mapping.keys() if col in df.columns]
    df = df[colunas_existentes]
    
    column_mapping = {k: v for k, v in column_mapping.items() if k in colunas_existentes}
    df = df.rename(columns=column_mapping)
    
    df['data_entrada'] = pd.to_datetime(df['data_entrada'], format='%d/%m/%Y', errors='coerce')
    df['total'] = df['total'].apply(clean_numeric)
    
    df = df.dropna(subset=['data_entrada'])
    df = df[df['total'] > 0]
    
    return df

@app.route('/')
def index():
    arquivos = [f for f in os.listdir(app.config['UPLOAD_FOLDER']) if f.endswith('.csv')]
    return render_template('index.html', arquivos=arquivos if arquivos else [])

@app.route('/upload', methods=['POST'])
def upload_file():
    if 'file' not in request.files:
        return jsonify({'error': 'Nenhum arquivo enviado'}), 400
    file = request.files['file']
    if file.filename == '':
        return jsonify({'error': 'Nenhum arquivo selecionado'}), 400
    if file and allowed_file(file.filename):
        filename = secure_filename(file.filename)
        file.save(os.path.join(app.config['UPLOAD_FOLDER'], filename))
        return jsonify({'success': True, 'filename': filename})
    return jsonify({'error': 'Tipo de arquivo não permitido'}), 400

@app.route('/delete_file', methods=['POST'])
def delete_file():
    filename = request.form.get('filename')
    if not filename:
        return jsonify({'error': 'Nome do arquivo não fornecido'}), 400
        
    filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
    try:
        if os.path.exists(filepath):
            os.remove(filepath)
            return jsonify({'success': True, 'message': f'Arquivo {filename} removido com sucesso'})
        return jsonify({'error': 'Arquivo não encontrado'}), 404
    except Exception as e:
        return jsonify({'error': f'Erro ao remover arquivo: {str(e)}'}), 500

@app.route('/api/data')
def get_data():
    try:
        arquivo = request.args.get('arquivo')
        if not arquivo:
            return jsonify({
                'summary': {},
                'table_data': [],
                'fornecedores': [],
                'clientes': [],
                'top_10_gastos': [],
                'analise_fornecedores': []
            })
            
        filepath = os.path.join(app.config['UPLOAD_FOLDER'], arquivo)
        if not os.path.exists(filepath):
            return jsonify({'error': 'Arquivo não encontrado'}), 404
            
        df = load_data(arquivo)
        fornecedor_filter = request.args.get('fornecedor')
        cliente_filter = request.args.get('cliente')
        
        if fornecedor_filter:
            df = df[df['fornecedor'].str.contains(fornecedor_filter, case=False, na=False)]
        if cliente_filter:
            df = df[df['cliente'].str.contains(cliente_filter, case=False, na=False)]
        
        fornecedores = sorted([f for f in df['fornecedor'].unique() if f and not pd.isna(f)])
        clientes = sorted([c for c in df['cliente'].unique() if c and not pd.isna(c)])
        
        # Análise geral
        summary = {
            'total_registros': int(len(df)),
            'total_valor': float(df['total'].sum()),
            'media_valor': float(df['total'].mean()),
            'total_por_tipo': {str(k): float(v) for k, v in df.groupby('tipo_custo')['total'].sum().items()},
            'evolucao_temporal': {str(k): float(v) for k, v in df.groupby(df['data_entrada'].dt.strftime('%Y-%m'))['total'].sum().items()}
        }
        
        # Top 10 maiores gastos
        top_10_gastos = df.nlargest(10, 'total').apply(lambda row: {
            'data': row['data_entrada'].strftime('%d/%m/%Y'),
            'fornecedor': str(row['fornecedor']),
            'cliente': str(row['cliente']),
            'obra': str(row['obra']),
            'tipo_custo': str(row['tipo_custo']),
            'valor': float(row['total'])
        }, axis=1).tolist()
        
        # Análise de fornecedores
        analise_fornecedores = []
        for nome, grupo in df.groupby('fornecedor'):
            if pd.isna(nome) or not nome:
                continue
                
            # Detalhes dos clientes e obras
            clientes_obras = {}
            for cliente in grupo['cliente'].unique():
                if pd.isna(cliente) or not cliente:
                    continue
                obras_cliente = grupo[grupo['cliente'] == cliente]['obra'].unique()
                obras_lista = [obra for obra in obras_cliente if not pd.isna(obra) and obra]
                clientes_obras[cliente] = obras_lista
            
            analise_fornecedores.append({
                'fornecedor': str(nome),
                'total_gasto': float(grupo['total'].sum()),
                'quantidade_transacoes': int(len(grupo)),
                'valor_medio': float(grupo['total'].mean()),
                'quantidade_clientes': int(grupo['cliente'].nunique()),
                'quantidade_obras': int(grupo['obra'].nunique()),
                'tipos_custo_diferentes': int(grupo['tipo_custo'].nunique()),
                'clientes_obras': clientes_obras
            })
        
        analise_fornecedores.sort(key=lambda x: x['total_gasto'], reverse=True)
        analise_fornecedores = analise_fornecedores[:20]
        
        # Dados da tabela
        table_data = []
        for _, row in df.iterrows():
            table_data.append({
                'data_entrada': row['data_entrada'].strftime('%d/%m/%Y'),
                'cliente': str(row['cliente']),
                'fornecedor': str(row['fornecedor']),
                'obra': str(row['obra']),
                'tipo_custo': str(row['tipo_custo']),
                'total': float(row['total'])
            })
        
        table_data = sorted(table_data, key=lambda x: x['data_entrada'], reverse=True)[:100]
        
        return jsonify({
            'summary': summary,
            'table_data': table_data,
            'fornecedores': fornecedores,
            'clientes': clientes,
            'top_10_gastos': top_10_gastos,
            'analise_fornecedores': analise_fornecedores
        })
        
    except Exception as e:
        print(f"Erro ao processar dados: {str(e)}")
        return jsonify({'error': str(e)}), 500

@app.route('/export/excel')
def export_excel():
    try:
        arquivo = request.args.get('arquivo')
        fornecedor = request.args.get('fornecedor')
        cliente = request.args.get('cliente')
        
        if not arquivo:
            return jsonify({'error': 'Arquivo não especificado'}), 400

        df = load_data(arquivo)
        
        if fornecedor:
            df = df[df['fornecedor'].str.contains(fornecedor, case=False, na=False)]
        if cliente:
            df = df[df['cliente'].str.contains(cliente, case=False, na=False)]

        # Criar workbook Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "Dados Exportados"

        # Estilo para cabeçalho
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")

        # Adicionar cabeçalhos
        headers = ['Data', 'Fornecedor', 'Cliente', 'Obra', 'Tipo Custo', 'Valor']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill

        # Adicionar dados
        for row_idx, row in enumerate(df.iterrows(), 2):
            ws.cell(row=row_idx, column=1, value=row[1]['data_entrada'].strftime('%d/%m/%Y'))
            ws.cell(row=row_idx, column=2, value=str(row[1]['fornecedor']))
            ws.cell(row=row_idx, column=3, value=str(row[1]['cliente']))
            ws.cell(row=row_idx, column=4, value=str(row[1]['obra']))
            ws.cell(row=row_idx, column=5, value=str(row[1]['tipo_custo']))
            ws.cell(row=row_idx, column=6, value=float(row[1]['total']))

        # Ajustar largura das colunas
        for column in ws.columns:
            max_length = 0
            column = [cell for cell in column]
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column[0].column_letter].width = adjusted_width

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return send_file(
            buffer,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'relatorio_filtrado_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        )    
    except Exception as e:
        return jsonify({'error': f'Erro no processamento: {str(e)}'}), 500 
    

if __name__ == '__main__':
    app.run(debug=True)

def export_to_csv(data):
    output = io.StringIO()
    df = pd.DataFrame(data)
    df.to_csv(output, index=False)
    output.seek(0)
    return output

@app.route('/export/<format>')
def export_data(format):
    arquivo = request.args.get('arquivo')
    fornecedor_filter = request.args.get('fornecedor')
    cliente_filter = request.args.get('cliente')
    export_type = request.args.get('type', 'table') # table, top10, fornecedores
    
    if not arquivo:
        return jsonify({'error': 'Arquivo não especificado'}), 400
        
    df = load_data(arquivo)
    
    if fornecedor_filter:
        df = df[df['fornecedor'].str.contains(fornecedor_filter, case=False, na=False)]
    if cliente_filter:
        df = df[df['cliente'].str.contains(cliente_filter, case=False, na=False)]
    
    if export_type == 'table':
        data = [{
            'Data': row['data_entrada'].strftime('%d/%m/%Y'),
            'Cliente': row['cliente'],
            'Fornecedor': row['fornecedor'],
            'Obra': row['obra'],
            'Tipo Custo': row['tipo_custo'],
            'Total': row['total']
        } for _, row in df.iterrows()]
    elif export_type == 'top10':
        data = df.nlargest(10, 'total').apply(lambda row: {
            'Data': row['data_entrada'].strftime('%d/%m/%Y'),
            'Fornecedor': row['fornecedor'],
            'Cliente': row['cliente'],
            'Obra': row['obra'],
            'Tipo Custo': row['tipo_custo'],
            'Valor': row['total']
        }, axis=1).tolist()
    elif export_type == 'fornecedores':
        data = [{
            'Fornecedor': nome,
            'Total Gasto': grupo['total'].sum(),
            'Qtd Transações': len(grupo),
            'Valor Médio': grupo['total'].mean(),
            'Qtd Clientes': grupo['cliente'].nunique(),
            'Qtd Obras': grupo['obra'].nunique()
        } for nome, grupo in df.groupby('fornecedor') if not pd.isna(nome) and nome]
       
    